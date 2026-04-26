"""
excel_export.py — FactuPro v2.0
Exportación profesional a Excel de reportes contables y DGII.
Usa openpyxl con formato corporativo.
"""

from io import BytesIO
from datetime import date, datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Colores corporativos FactuPro ─────────────────────────────────────────────
COLOR_HEADER_BG  = "0F1C2E"   # Azul oscuro navy
COLOR_HEADER_FG  = "FFFFFF"   # Blanco
COLOR_ACCENT     = "F97316"   # Naranja FactuPro
COLOR_ACCENT_FG  = "FFFFFF"
COLOR_SUBHEADER  = "162236"   # Navy2
COLOR_ALT_ROW    = "F0F4F9"   # Gris claro
COLOR_TOTAL_BG   = "1E304A"   # Navy3
COLOR_TOTAL_FG   = "FFFFFF"
COLOR_GREEN_BG   = "D1FAE5"
COLOR_RED_BG     = "FEE2E2"
COLOR_BORDER     = "CBD5E1"

# Formato moneda DOP
FMT_CURRENCY = '#,##0.00'
FMT_DATE     = 'DD/MM/YYYY'


def _border(style="thin"):
    s = Side(style=style, color=COLOR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, row, col, value, bold=True, size=11, bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border    = _border()
    return cell


def _data_style(ws, row, col, value, fmt=None, bold=False, bg=None, fg="000000", align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _border("thin")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    return cell


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _add_logo_header(ws, title: str, empresa: dict, desde: str, hasta: str):
    """Cabecera corporativa con datos de la empresa."""
    # Fila 1 — Título principal
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "FactuPro — Sistema de Facturación"
    c.font      = Font(name="Arial", bold=True, size=14, color=COLOR_HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2 — Nombre empresa
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = empresa.get("nombre", "Mi Empresa")
    c.font      = Font(name="Arial", bold=True, size=12, color=COLOR_ACCENT)
    c.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Fila 3 — Título del reporte
    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value     = title
    c.font      = Font(name="Arial", bold=True, size=11, color=COLOR_HEADER_FG)
    c.fill      = PatternFill("solid", fgColor=COLOR_SUBHEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 20

    # Fila 4 — Período y RNC
    ws.merge_cells("A4:D4")
    c = ws["A4"]
    c.value     = f"Período: {desde}  al  {hasta}"
    c.font      = Font(name="Arial", size=10, color="374151")
    c.fill      = PatternFill("solid", fgColor="F8FAFC")
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("E4:H4")
    c = ws["E4"]
    rnc = empresa.get("rnc") or "—"
    c.value     = f"RNC: {rnc}  |  Generado: {date.today().strftime('%d/%m/%Y')}"
    c.font      = Font(name="Arial", size=10, color="374151")
    c.fill      = PatternFill("solid", fgColor="F8FAFC")
    c.alignment = Alignment(horizontal="right", vertical="center")

    ws.row_dimensions[4].height = 18

    return 5   # próxima fila disponible


# ════════════════════════════════════════════════════════════════════════════
# REPORTE CONTABLE — Transacciones
# ════════════════════════════════════════════════════════════════════════════

def exportar_contabilidad(transacciones: list, empresa: dict,
                          desde: str, hasta: str) -> bytes:
    """
    Genera Excel de transacciones contables con:
    - Hoja 1: Detalle de transacciones
    - Hoja 2: Resumen por categoría
    """
    wb = Workbook()

    # ── Hoja 1: Detalle ───────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Transacciones"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "A7"

    fila = _add_logo_header(ws1, "Registro de Transacciones Contables",
                            empresa, desde, hasta)

    # Fila vacía
    fila += 1

    # Encabezados
    headers = ["Fecha", "Tipo", "Categoría", "Descripción",
               "Método", "Referencia", "Monto (RD$)"]
    widths  = [14, 12, 20, 40, 16, 20, 18]
    _set_col_widths(ws1, widths)
    ws1.row_dimensions[fila].height = 22

    for col, h in enumerate(headers, 1):
        _header_style(ws1, fila, col, h, bg=COLOR_ACCENT, fg=COLOR_HEADER_FG)

    fila += 1

    # Datos
    total_ing = 0.0
    total_egr = 0.0

    for i, t in enumerate(transacciones):
        bg = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        is_ing = t.get("tipo","").upper() == "INGRESO"
        monto  = float(t.get("monto", 0))

        _data_style(ws1, fila, 1, t.get("fecha", ""), fmt=FMT_DATE, bg=bg, align="center")
        # Tipo con color
        tipo_cell = _data_style(ws1, fila, 2, t.get("tipo",""), bold=True,
                                bg=COLOR_GREEN_BG if is_ing else COLOR_RED_BG,
                                fg="065F46" if is_ing else "991B1B", align="center")
        _data_style(ws1, fila, 3, t.get("categoria","—"), bg=bg)
        _data_style(ws1, fila, 4, t.get("descripcion",""), bg=bg)
        _data_style(ws1, fila, 5, t.get("metodo",""), bg=bg, align="center")
        _data_style(ws1, fila, 6, t.get("referencia","—"), bg=bg, align="center")
        monto_cell = _data_style(ws1, fila, 7, monto, fmt=FMT_CURRENCY,
                                 bold=True, bg=bg, align="right",
                                 fg="065F46" if is_ing else "991B1B")

        if is_ing: total_ing += monto
        else:      total_egr += monto

        ws1.row_dimensions[fila].height = 16
        fila += 1

    # Fila de totales
    fila += 1
    ws1.merge_cells(f"A{fila}:F{fila}")
    c = ws1.cell(row=fila, column=1, value="TOTAL INGRESOS")
    c.font      = Font(name="Arial", bold=True, color=COLOR_TOTAL_FG, size=10)
    c.fill      = PatternFill("solid", fgColor="065F46")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws1.cell(row=fila, column=7, value=total_ing)
    c.font            = Font(name="Arial", bold=True, color=COLOR_TOTAL_FG, size=11)
    c.fill            = PatternFill("solid", fgColor="065F46")
    c.alignment       = Alignment(horizontal="right", vertical="center")
    c.number_format   = FMT_CURRENCY
    ws1.row_dimensions[fila].height = 20

    fila += 1
    ws1.merge_cells(f"A{fila}:F{fila}")
    c = ws1.cell(row=fila, column=1, value="TOTAL EGRESOS")
    c.font      = Font(name="Arial", bold=True, color=COLOR_TOTAL_FG, size=10)
    c.fill      = PatternFill("solid", fgColor="991B1B")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws1.cell(row=fila, column=7, value=total_egr)
    c.font            = Font(name="Arial", bold=True, color=COLOR_TOTAL_FG, size=11)
    c.fill            = PatternFill("solid", fgColor="991B1B")
    c.alignment       = Alignment(horizontal="right", vertical="center")
    c.number_format   = FMT_CURRENCY
    ws1.row_dimensions[fila].height = 20

    fila += 1
    utilidad = total_ing - total_egr
    ws1.merge_cells(f"A{fila}:F{fila}")
    c = ws1.cell(row=fila, column=1, value="UTILIDAD NETA")
    c.font      = Font(name="Arial", bold=True, color=COLOR_TOTAL_FG, size=11)
    c.fill      = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c = ws1.cell(row=fila, column=7, value=utilidad)
    c.font            = Font(name="Arial", bold=True, color=COLOR_ACCENT, size=12)
    c.fill            = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    c.alignment       = Alignment(horizontal="right", vertical="center")
    c.number_format   = FMT_CURRENCY
    ws1.row_dimensions[fila].height = 22

    # ── Hoja 2: Resumen por categoría ─────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen por Categoría")
    ws2.sheet_view.showGridLines = False

    fila2 = _add_logo_header(ws2, "Resumen Contable por Categoría",
                             empresa, desde, hasta)
    fila2 += 1
    _set_col_widths(ws2, [30, 18, 18, 18])

    headers2 = ["Categoría", "Ingresos (RD$)", "Egresos (RD$)", "Neto (RD$)"]
    ws2.row_dimensions[fila2].height = 22
    for col, h in enumerate(headers2, 1):
        _header_style(ws2, fila2, col, h, bg=COLOR_ACCENT)
    fila2 += 1

    # Agrupar por categoría
    cats = {}
    for t in transacciones:
        cat   = t.get("categoria") or "Sin categoría"
        tipo  = t.get("tipo","").upper()
        monto = float(t.get("monto", 0))
        if cat not in cats:
            cats[cat] = {"ing": 0, "egr": 0}
        if tipo == "INGRESO": cats[cat]["ing"] += monto
        else:                 cats[cat]["egr"] += monto

    for i, (cat, vals) in enumerate(sorted(cats.items())):
        bg    = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        neto  = vals["ing"] - vals["egr"]
        neto_bg = COLOR_GREEN_BG if neto >= 0 else COLOR_RED_BG
        neto_fg = "065F46" if neto >= 0 else "991B1B"
        _data_style(ws2, fila2, 1, cat, bg=bg, bold=True)
        _data_style(ws2, fila2, 2, vals["ing"], fmt=FMT_CURRENCY, bg=bg, align="right", fg="065F46")
        _data_style(ws2, fila2, 3, vals["egr"], fmt=FMT_CURRENCY, bg=bg, align="right", fg="991B1B")
        _data_style(ws2, fila2, 4, neto,        fmt=FMT_CURRENCY, bg=neto_bg, align="right", fg=neto_fg, bold=True)
        ws2.row_dimensions[fila2].height = 16
        fila2 += 1

    # Totales
    fila2 += 1
    _header_style(ws2, fila2, 1, "TOTALES", bg=COLOR_HEADER_BG, align="right")
    _header_style(ws2, fila2, 2, total_ing, bg="065F46")
    ws2.cell(row=fila2, column=2).number_format = FMT_CURRENCY
    _header_style(ws2, fila2, 3, total_egr, bg="991B1B")
    ws2.cell(row=fila2, column=3).number_format = FMT_CURRENCY
    col_net = _header_style(ws2, fila2, 4, total_ing-total_egr,
                            bg=COLOR_ACCENT if (total_ing-total_egr) >= 0 else "991B1B")
    ws2.cell(row=fila2, column=4).number_format = FMT_CURRENCY
    ws2.row_dimensions[fila2].height = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# REPORTE 606 — Compras (DGII)
# ════════════════════════════════════════════════════════════════════════════

def exportar_606(filas: list, empresa: dict, desde: str, hasta: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "606 - Compras DGII"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    fila = _add_logo_header(ws, "Formato 606 — Registro de Compras (DGII)",
                            empresa, desde, hasta)
    fila += 1

    headers = ["RNC Proveedor", "Nombre Proveedor", "Tipo Bienes/Servicios",
               "Fecha", "Número", "Monto (RD$)", "ITBIS (RD$)", "Total (RD$)"]
    widths  = [18, 35, 22, 14, 18, 18, 18, 18]
    _set_col_widths(ws, widths)
    ws.row_dimensions[fila].height = 22

    for col, h in enumerate(headers, 1):
        _header_style(ws, fila, col, h, bg=COLOR_ACCENT)
    fila += 1

    total_monto = total_itbis = total_total = 0.0

    for i, f in enumerate(filas):
        bg     = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        monto  = float(f.get("monto", 0))
        itbis  = float(f.get("itbis", 0))
        total  = monto + itbis
        _data_style(ws, fila, 1, f.get("rnc_proveedor","—"), bg=bg, align="center")
        _data_style(ws, fila, 2, f.get("nombre","—"), bg=bg)
        _data_style(ws, fila, 3, "01", bg=bg, align="center")  # Bienes
        _data_style(ws, fila, 4, f.get("fecha",""), bg=bg, align="center")
        _data_style(ws, fila, 5, f.get("numero","—"), bg=bg, align="center")
        _data_style(ws, fila, 6, monto, fmt=FMT_CURRENCY, bg=bg, align="right")
        _data_style(ws, fila, 7, itbis, fmt=FMT_CURRENCY, bg=bg, align="right")
        _data_style(ws, fila, 8, total, fmt=FMT_CURRENCY, bg=bg, align="right", bold=True)
        ws.row_dimensions[fila].height = 16
        total_monto += monto; total_itbis += itbis; total_total += total
        fila += 1

    # Totales
    fila += 1
    for col, (val, label) in enumerate([
        (None,"TOTALES"), (None,""), (None,""), (None,""),
        (None,""), (total_monto,None), (total_itbis,None), (total_total,None)
    ], 1):
        if label is not None:
            c = _header_style(ws, fila, col, label, bg=COLOR_HEADER_BG, align="right")
        else:
            c = _header_style(ws, fila, col, val, bg=COLOR_HEADER_BG, align="right")
            c.number_format = FMT_CURRENCY
    ws.row_dimensions[fila].height = 22

    # Nota legal
    fila += 2
    ws.merge_cells(f"A{fila}:H{fila}")
    c = ws.cell(row=fila, column=1,
                value="Nota: Este reporte corresponde al Formulario IR-17 (Formato 606) de la DGII, República Dominicana.")
    c.font = Font(name="Arial", italic=True, size=9, color="6B7280")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# REPORTE 607 — Ventas (DGII)
# ════════════════════════════════════════════════════════════════════════════

def exportar_607(filas: list, empresa: dict, desde: str, hasta: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "607 - Ventas DGII"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A7"

    fila = _add_logo_header(ws, "Formato 607 — Registro de Ventas (DGII)",
                            empresa, desde, hasta)
    fila += 1

    headers = ["RNC Cliente", "Nombre Cliente", "Tipo ID",
               "NCF", "Tipo NCF", "Fecha", "Subtotal (RD$)",
               "ITBIS (RD$)", "Total (RD$)"]
    widths  = [18, 32, 12, 18, 10, 14, 18, 18, 18]
    _set_col_widths(ws, widths)
    ws.row_dimensions[fila].height = 22

    for col, h in enumerate(headers, 1):
        _header_style(ws, fila, col, h, bg=COLOR_ACCENT)
    fila += 1

    total_sub = total_itbis = total_tot = 0.0

    for i, f in enumerate(filas):
        bg      = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        sub     = float(f.get("subtotal", 0))
        itbis   = float(f.get("itbis", 0))
        total   = float(f.get("total", 0))
        _data_style(ws, fila, 1, f.get("rnc_cliente","—"), bg=bg, align="center")
        _data_style(ws, fila, 2, f.get("nombre","—"), bg=bg)
        _data_style(ws, fila, 3, f.get("tipo_id","FINAL"), bg=bg, align="center")
        _data_style(ws, fila, 4, f.get("ncf","—"), bg=bg, align="center")
        _data_style(ws, fila, 5, f.get("tipo_ncf","B02"), bg=bg, align="center")
        _data_style(ws, fila, 6, f.get("fecha",""), bg=bg, align="center")
        _data_style(ws, fila, 7, sub,   fmt=FMT_CURRENCY, bg=bg, align="right")
        _data_style(ws, fila, 8, itbis, fmt=FMT_CURRENCY, bg=bg, align="right")
        _data_style(ws, fila, 9, total, fmt=FMT_CURRENCY, bg=bg, align="right", bold=True)
        ws.row_dimensions[fila].height = 16
        total_sub += sub; total_itbis += itbis; total_tot += total
        fila += 1

    # Totales
    fila += 1
    ws.merge_cells(f"A{fila}:F{fila}")
    c = _header_style(ws, fila, 1, "TOTALES", bg=COLOR_HEADER_BG, align="right")
    _header_style(ws, fila, 7, total_sub,   bg=COLOR_HEADER_BG).number_format = FMT_CURRENCY
    _header_style(ws, fila, 8, total_itbis, bg=COLOR_HEADER_BG).number_format = FMT_CURRENCY
    _header_style(ws, fila, 9, total_tot,   bg=COLOR_ACCENT   ).number_format = FMT_CURRENCY
    ws.row_dimensions[fila].height = 22

    # Nota legal
    fila += 2
    ws.merge_cells(f"A{fila}:I{fila}")
    c = ws.cell(row=fila, column=1,
                value="Nota: Este reporte corresponde al Formulario IR-17 (Formato 607) de la DGII, República Dominicana.")
    c.font = Font(name="Arial", italic=True, size=9, color="6B7280")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ════════════════════════════════════════════════════════════════════════════
# REPORTE VENTAS GENERAL
# ════════════════════════════════════════════════════════════════════════════

def exportar_reporte_ventas(detalle: list, resumen: dict,
                            empresa: dict, desde: str, hasta: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A9"

    fila = _add_logo_header(ws, "Reporte de Ventas", empresa, desde, hasta)

    # KPIs
    fila += 1
    kpis = [
        ("Total Ventas",  resumen.get("total_ventas",0),  COLOR_HEADER_BG, COLOR_HEADER_FG),
        ("Total Cobrado", resumen.get("total_cobrado",0), "065F46",        "FFFFFF"),
        ("Por Cobrar",    resumen.get("total_balance",0), "991B1B",        "FFFFFF"),
        ("# Facturas",    resumen.get("total_facturas",0),"1E40AF",        "FFFFFF"),
    ]
    for col, (label, val, bg, fg) in enumerate(kpis, 1):
        ws.merge_cells(start_row=fila, start_column=(col-1)*2+1,
                       end_row=fila, end_column=(col-1)*2+2)
        c = ws.cell(row=fila, column=(col-1)*2+1, value=label)
        c.font = Font(name="Arial", bold=True, color=fg, size=10)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = 18
        fila += 1
        ws.merge_cells(start_row=fila, start_column=(col-1)*2+1,
                       end_row=fila, end_column=(col-1)*2+2)
        c = ws.cell(row=fila, column=(col-1)*2+1, value=val)
        c.font = Font(name="Arial", bold=True, color=fg, size=14)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col < 4: c.number_format = FMT_CURRENCY
        ws.row_dimensions[fila].height = 26
        fila -= 1

    fila += 3

    # Encabezados
    headers = ["NCF", "Cliente", "Fecha", "Total (RD$)",
               "Cobrado (RD$)", "Balance (RD$)", "Estatus"]
    widths  = [18, 35, 14, 18, 18, 18, 14]
    _set_col_widths(ws, widths)
    ws.row_dimensions[fila].height = 22
    for col, h in enumerate(headers, 1):
        _header_style(ws, fila, col, h, bg=COLOR_ACCENT)
    fila += 1

    for i, f in enumerate(detalle):
        bg      = COLOR_ALT_ROW if i % 2 == 0 else "FFFFFF"
        balance = float(f.get("balance", 0))
        bal_bg  = COLOR_RED_BG if balance > 0 else COLOR_GREEN_BG
        bal_fg  = "991B1B" if balance > 0 else "065F46"
        _data_style(ws, fila, 1, f.get("ncf","—"), bg=bg, align="center")
        _data_style(ws, fila, 2, f.get("cliente","—"), bg=bg)
        _data_style(ws, fila, 3, f.get("fecha",""), bg=bg, align="center")
        _data_style(ws, fila, 4, float(f.get("total",0)),   fmt=FMT_CURRENCY, bg=bg, align="right", bold=True)
        _data_style(ws, fila, 5, float(f.get("cobrado",0)), fmt=FMT_CURRENCY, bg=bg, align="right", fg="065F46")
        _data_style(ws, fila, 6, balance,                   fmt=FMT_CURRENCY, bg=bal_bg, align="right", fg=bal_fg, bold=True)
        _data_style(ws, fila, 7, f.get("estatus","—"), bg=bg, align="center",
                    bold=True, fg="065F46" if f.get("estatus") == "PAGADA" else "991B1B")
        ws.row_dimensions[fila].height = 16
        fila += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
